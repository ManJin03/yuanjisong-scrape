# -*- coding: utf-8 -*-
"""Excel 导出：多 Sheet、表头样式、自适应列宽、冻结首行、职位超链接。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from yuanjisong.models import Project
from yuanjisong import config

COLUMNS = [
    ("ID", lambda p: p.id),
    ("标题", lambda p: p.title),
    ("预算(元)", lambda p: p.budget),
    ("工时", lambda p: f"{p.hours:g} {p.hours_unit}".strip()),
    ("类型", lambda p: p.work_type),
    ("远程", lambda p: "是" if p.is_remote else "否"),
    ("已投递", lambda p: p.delivery_count),
    ("技术分类", lambda p: p.category or "未分类"),
    ("黑名单", lambda p: p.blacklist_hit or "-"),
    ("命中词", lambda p: p.blacklist_word or "-"),
    ("雇主", lambda p: p.employer_name),
    ("描述", lambda p: p.description),
    ("链接", lambda p: p.url),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _write_sheet(ws, projects: list[Project], link_col: str | None = "链接") -> None:
    ws.append([name for name, _ in COLUMNS])
    for cell in ws[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for p in projects:
        ws.append([fn(p) for _, fn in COLUMNS])
    # 自适应列宽（描述限宽）
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row):
            w = min(max(widths.get(i, 8), len(str(v)) * 1.8 if v else 8), 60)
            widths[i] = w
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i + 1)].width = min(w, 60)
    ws.freeze_panes = "A2"
    if link_col:
        idx = next(i for i, (n, _) in enumerate(COLUMNS) if n == link_col) + 1
        letter = get_column_letter(idx)
        for r in range(2, ws.max_row + 1):
            cell = ws[f"{letter}{r}"]
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")


def export_all_excel(projects: list[Project], path=None) -> Path:
    """导出全量项目 + 按技术方向分 Sheet。"""
    path = Path(path or config.EXCEL_ALL)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_sheet(wb.active, projects)
    wb.active.title = "全部项目"
    for sheet in config.ALL_SHEETS:
        items = [p for p in projects if (p.category or config.CATEGORY_OTHER) == sheet]
        if items:
            _write_sheet(wb.create_sheet(sheet[:31]), items)
    wb.save(path)
    return path


def export_student_excel(projects: list[Project], path=None) -> Path:
    """导出学生友好清单（调用前应已完成筛选与排序）。"""
    path = Path(path or config.EXCEL_STUDENT)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_sheet(wb.active, projects)
    wb.active.title = f"学生友好(≤{config.STUDENT_MAX_BUDGET}元·远程)"
    wb.save(path)
    return path
