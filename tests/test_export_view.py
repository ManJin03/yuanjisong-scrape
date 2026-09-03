# -*- coding: utf-8 -*-
"""export_excel（当前视图导出）测试。"""
from pathlib import Path

from openpyxl import load_workbook

from yuanjisong.exporter import export_excel
from yuanjisong.models import parse_job_cards

FIXTURE = Path(__file__).parent / "fixtures" / "two_cards.html"


def test_export_view(tmp_path: Path):
    ps = parse_job_cards(FIXTURE.read_text(encoding="utf-8"))
    path = export_excel(ps, tmp_path / "view.xlsx", sheet_name="当前视图")
    wb = load_workbook(path)
    ws = wb["当前视图"]
    assert ws.max_row == 1 + len(ps)
    assert ws.cell(row=1, column=1).value == "ID"
    assert ws.cell(row=2, column=2).value == ps[0].title
