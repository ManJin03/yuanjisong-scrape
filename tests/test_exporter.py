# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import load_workbook

from yuanjisong.classify import classify_all
from yuanjisong.exporter import export_all_excel, export_student_excel
from yuanjisong.filter_student_projects import filter_student_projects
from yuanjisong.models import parse_job_cards


def load_fixture_projects():
    html = (Path(__file__).parent / "fixtures" / "job_page1.html").read_text(encoding="utf-8")
    return parse_job_cards(html, page=1)


def test_export_all(tmp_path: Path):
    ps = classify_all(load_fixture_projects())
    path = export_all_excel(ps, tmp_path / "all.xlsx")
    wb = load_workbook(path)
    assert "全部项目" in wb.sheetnames
    assert wb["全部项目"].max_row == 1 + len(ps)
    assert any("爬虫" in name or "前端" in name or "后端" in name for name in wb.sheetnames)
    ws = wb["全部项目"]
    assert ws.cell(row=1, column=1).value == "ID"
    assert ws.cell(row=2, column=3).value > 0  # 预算数值


def test_export_student(tmp_path: Path):
    ps = filter_student_projects(load_fixture_projects())
    path = export_student_excel(ps, tmp_path / "student.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row >= 1
    budgets = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    assert all(0 < b <= 500 for b in budgets)
    assert budgets == sorted(budgets)
